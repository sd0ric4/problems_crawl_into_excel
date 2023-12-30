import json
import openpyxl
import os

# JSON 文件路径
json_file_path = './json/'
# Excel 文件保存路径
excel_file_path = './excel/'

# 遍历 JSON 文件夹中的所有文件
for root, dirs, files in os.walk(json_file_path):
    for file in files:
        if file.endswith('.json'):
            json_file = os.path.join(root, file)
            # 读取 JSON 文件
            with open(json_file, 'r') as json_data_file:
                json_data = json.load(json_data_file)

            # 创建 Excel 工作簿和工作表
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            # 添加表头
            worksheet['A1'] = '大题题干'
            worksheet['B1'] = '题型(选填)'
            worksheet['C1'] = '题干/小题题干'
            worksheet['D1'] = '选项A'
            worksheet['E1'] = '选项B'
            worksheet['F1'] = '选项C'
            worksheet['G1'] = '选项D'
            worksheet['H1'] = '选项E'
            worksheet['I1'] = '答案'
            worksheet['J1'] = '解析'
            worksheet['K1'] = '章节'
            worksheet['L1'] = '知识点'
            worksheet['M1'] = '分值'

            # 遍历 JSON 数据并写入 Excel 表格
            for row_index, data in enumerate(json_data, start=2):
                chapter = data['chapter']
                type = data['type']
                title = data['title']
                options = {list(option.keys())[0]: list(option.values())[0] for option in data['options']}
                correct_answer = data['correctAnswer']
                worksheet.cell(row=row_index, column=2, value=type)  # 题型
                worksheet.cell(row=row_index, column=3, value=title)  # 题干
                worksheet.cell(row=row_index, column=4, value=options.get('A', ''))  # 选项A
                worksheet.cell(row=row_index, column=5, value=options.get('B', ''))  # 选项B
                worksheet.cell(row=row_index, column=6, value=options.get('C', ''))  # 选项C
                worksheet.cell(row=row_index, column=7, value=options.get('D', ''))  # 选项D
                worksheet.cell(row=row_index, column=8, value=options.get('E', ''))  # 选项E
                worksheet.cell(row=row_index, column=9, value=correct_answer)  # 答案
                worksheet.cell(row=row_index, column=11, value=chapter)  # 章节

            # 保存 Excel 文件
            excel_file = os.path.join(excel_file_path, f"{os.path.splitext(file)[0]}.xlsx")
            workbook.save(excel_file)